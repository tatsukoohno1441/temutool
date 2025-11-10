#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_report_v5.py  (Case‑Insensitive header fix)
------------------------------------------------
* 追加した Order item ID 認識が **大文字小文字・空白** の違いでエラーになる問題を修正。
  → ヘッダーは `strip().lower()` で全列を辞書化し、大小文字を無視してマッチ。
* そのほか仕様（qty=quantity to ship, 新列追加など）は v5 と同じ。
"""
import sys, os
import pandas as pd

# --- 日英混合ヘッダー対応 --------------------------------------------------
COL_ALIAS = {
    # 订单编号
    "order id":                       "order id",
    "注文id":                          "order id",

    # 订单明细编号
    "order item id":                  "order item id",
    "注文商品id":                       "order item id",

    # 商品名
    "product name by customer order": "product name by customer order",
    "顧客注文による製品名":                "product name by customer order",

    # SKU
    "contribution sku":               "contribution sku",
    "貢献sku":                         "contribution sku",

    # 数量
    "quantity to ship":               "quantity to ship",
    "出荷数量":                         "quantity to ship",

    # 收件人
    "recipient name":                 "recipient name",
    "受取人名":                         "recipient name",
}

REQ_LOWER = [
    "order id",
    "order item id",
    "recipient name",
    "contribution sku",
    "product name by customer order",
    "quantity to ship",
]

HEAD_MAP = {
    "order id": "order_id",
    "order item id": "order_item_id",
    "recipient name": "recipient",
    "contribution sku": "jan",
    "product name by customer order": "product",
    "quantity to ship": "qty",
}


def read_table(path: str) -> pd.DataFrame:
    # ❶ 读文件（全部按字符串）
    if path.lower().endswith((".xlsx", ".xlsm", ".xls")):
        df = pd.read_excel(path, dtype=str, engine="openpyxl")
    else:
        df = pd.read_csv(path, dtype=str, keep_default_na=False)

    # ★★★ ❷ 第一次列名统一：日文 → 英文长名（小写） ★★★
    df.rename(
        columns=lambda c: COL_ALIAS.get(c.strip().lower(), c.strip().lower()),
        inplace=True
    )

    # ❸ 用 lower -> original 映射表做必需列检查
    col_map = {c.strip().lower(): c for c in df.columns}

    miss = [h for h in REQ_LOWER if h not in col_map]
    if miss:
        raise SystemExit(f"Missing header(s): {miss}\\nFound: {list(df.columns)}")

    # ❹ 第二次列名统一：英文长名 → 内部短名
    rename_dict = {col_map[k]: HEAD_MAP[k] for k in REQ_LOWER}
    df = df.rename(columns=rename_dict)

    # ❺ qty → int
    q = (
        df["qty"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.replace(",", "", regex=False)
    )
    df["qty"] = pd.to_numeric(q, errors="coerce").fillna(0).astype(int)

    # ❻ 其余列转 str
    for c in ["order_id", "order_item_id", "jan", "recipient", "product"]:
        df[c] = df[c].fillna("").astype(str)

    return df

# ---- build_detail, jan_totals, write_excel, main は変更なし ↓ ----

def build_detail(df: pd.DataFrame) -> pd.DataFrame:
    detail = df.groupby(["jan","order_id","order_item_id","recipient","product"], as_index=False)["qty"].sum()
    detail["global_lines"] = detail.groupby("order_id")["product"].transform("count")
    detail["jan_lines"] = detail.groupby(["jan","order_id"])["product"].transform("count")
    return detail.sort_values(["jan","order_id","order_item_id","recipient","product"], kind="stable")

def jan_totals(df: pd.DataFrame) -> pd.DataFrame:
    return (df.groupby(["jan","product"], as_index=False)["qty"].sum()
              .sort_values(["jan","qty","product"], ascending=[True,False,True], kind="stable"))

def write_excel(detail: pd.DataFrame, jan_total: pd.DataFrame, outp: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule

    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("整理結果")
    ws.append(["受注番号/Order ID","Order item ID","宛先名/Recipient","JANコード/JAN","商品名/Product","出荷個数/Qty to ship"])
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="F2F2F2")

    for idx, jan in enumerate(detail["jan"].drop_duplicates()):
        singles = detail[(detail["jan"]==jan) & (detail["jan_lines"]==1) & (detail["global_lines"]==1)]
        for _, r in singles.iterrows():
            ws.append([r["order_id"], r["order_item_id"], r["recipient"], r["jan"], r["product"], int(r["qty"])])
        if idx != len(detail["jan"].drop_duplicates())-1:
            ws.append(["","","","","",""])

    if ws.max_row>=2:
        ws.conditional_formatting.add(f"A2:F{ws.max_row}",
            FormulaRule(formula=["$F2>1"], font=Font(bold=True,color="FF0000")))

    multi_ids = detail.loc[detail["global_lines"]>=2, "order_id"].drop_duplicates().tolist()
    if multi_ids:
        ws.append(["","","","","",""])
        ws.append(["-- 複数行の注文（全体） --","","","","",""])
        for c in ws[ws.max_row]: c.font = Font(bold=True)
        multis = detail[detail["order_id"].isin(multi_ids)]
        multis = multis.sort_values(["order_id","order_item_id","jan","recipient","product"], kind="stable")
        current=None
        for _, r in multis.iterrows():
            if current!=r["order_id"]:
                if current is not None:
                    ws.append(["","","","","",""])
                current=r["order_id"]
            ws.append([r["order_id"], r["order_item_id"], r["recipient"], r["jan"], r["product"], int(r["qty"])])

    for col,w in zip(["A","B","C","D","E","F"],[18,18,20,18,36,10]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("JAN合計")
    ws2.append(["JANコード","商品名","合計数量"])
    for c in ws2[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="F2F2F2")
    for _, r in jan_total.iterrows():
        ws2.append([r["jan"], r["product"], int(r["qty"])])
    for col,w in zip(["A","B","C"],[18,36,12]):
        ws2.column_dimensions[col].width=w

    wb.save(outp)


def main():
    if len(sys.argv)<2:
        print("Usage: python3 build_report_v5.py input.xlsx [output.xlsx]"); sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv)>=3 else os.path.splitext(inp)[0]+"_report.xlsx"
    df = read_table(inp)
    detail = build_detail(df)
    jan_total = jan_totals(df)
    write_excel(detail, jan_total, out)
    print("Done:", out)

if __name__ == "__main__":
    main()
