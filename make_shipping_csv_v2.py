#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_shipping_csv_v2.py

Create a shipping CSV that:
- Keeps the SAME column order/names as the original CSV.
- Row order follows the Order ID sequence in report.xlsx (sheet "整理結果").
- Normalizes the "recipient phone number" column: remove '+81' and replace spaces with '-'.
- If "district" is blank, copy value from "ship address 1".
- Deduplicate multi-line orders (same Order ID): keep only first occurrence after sorting.
- Also output an XLSX "shipping_formatted.xlsx" with those multi-line rows colored blue.

Usage:
    python3 make_shipping_csv_v2.py original.csv report.xlsx shipping.csv
"""
import sys, re
import pandas as pd

COL_ALIAS = {
    "注文id":            "order id",
    "order id":         "order id",

    "注文商品id":         "order item id",
    "order item id":    "order item id",

    "受取人名":           "recipient name",
    "recipient name":   "recipient name",

    "貢献sku":           "contribution sku",
    "contribution sku": "contribution sku",

    "顧客注文による製品名": "product name by customer order",
    "product name by customer order": "product name by customer order",

    "出荷数量":           "quantity to ship",
    "quantity to ship": "quantity to ship",

    "受信者の電話番号": "recipient phone number",
    "地区": "district",
    "発送先住所1": "ship address 1",
    
}


def read_csv_all_text(path):
    df = pd.read_csv(path, dtype=str, keep_default_na=False)

    #日文・英語ヘッダー統一
    df.rename(
        columns=lambda c: COL_ALIAS.get(c.strip().lower(), c.strip().lower()),
        inplace=True
    )

    return df

def extract_order_sequence(report_xlsx):
    df = pd.read_excel(report_xlsx, sheet_name="整理結果", dtype=str, engine="openpyxl")
    col = None
    if "受注番号/Order ID" in df.columns:
        col = "受注番号/Order ID"
    else:
        cand = [c for c in df.columns if isinstance(c, str) and "Order ID" in c]
        if cand:
            col = cand[0]
    if not col:
        raise SystemExit("Could not find Order ID column in 整理結果.")
    return [str(x).strip() for x in df[col].tolist() if str(x).strip() != ""]

def normalize_phone(val: str) -> str:
    if val is None:
        return val
    t = str(val)
    t = t.replace("+81", "")
    t = re.sub(r"\s+", "-", t.strip())
    return t

def find_col_case_insensitive(df, name):
    for c in df.columns:
        if isinstance(c, str) and c.lower().strip() == name.lower().strip():
            return c
    return None

def main():
    if len(sys.argv) < 4:
        print("Usage: python3 make_shipping_csv_v2.py original.csv report.xlsx shipping.csv")
        sys.exit(1)
    src_csv, report_xlsx, out_csv = sys.argv[1], sys.argv[2], sys.argv[3]

    orig = read_csv_all_text(src_csv)

    # Ensure order id column
    if "order id" not in orig.columns:
        oc = find_col_case_insensitive(orig, "order id")
        if oc is None:
            raise SystemExit("Original CSV missing 'order id' column.")
        if oc != "order id":
            orig.rename(columns={oc:"order id"}, inplace=True)

    # Phone normalization
    phone_col = find_col_case_insensitive(orig, "recipient phone number")
    if phone_col:
        orig[phone_col] = orig[phone_col].map(normalize_phone)

    # District fill from Ship Address 1 when blank
    district_col = find_col_case_insensitive(orig, "district")
    addr1_col   = find_col_case_insensitive(orig, "ship address 1")
    if district_col and addr1_col:
        mask_blank = orig[district_col].astype(str).str.strip().eq("")
        orig.loc[mask_blank, district_col] = orig.loc[mask_blank, addr1_col]

    # Desired order from report
    order_seq = extract_order_sequence(report_xlsx)
    order_set = set(orig["order id"].astype(str))
    seq_filtered = [oid for oid in order_seq if oid in order_set]

    # Stable sort by sequence
    order_rank = {oid: i for i, oid in enumerate(seq_filtered)}
    orig["_rank"] = orig["order id"].astype(str).map(lambda x: order_rank.get(x, 10**12))
    orig["_idx"] = range(len(orig))
    sorted_df = orig.sort_values(by=["_rank","_idx"], kind="stable")

    # Deduplicate multi-line orders after sort (keep first occurrence)
    counts = sorted_df["order id"].astype(str).value_counts()
    multi_ids = set(counts[counts > 1].index)
    keep_list = []
    seen = set()
    for oid in sorted_df["order id"].astype(str).tolist():
        if oid in multi_ids:
            if oid in seen:
                keep_list.append(False)
            else:
                keep_list.append(True)
                seen.add(oid)
        else:
            keep_list.append(True)

    sorted_df = sorted_df.loc[keep_list].copy()
    sorted_df.drop(columns=["_rank","_idx"], inplace=True)

    # Write CSV (UTF-8 BOM for Japanese)
    sorted_df.to_csv(out_csv, index=False, encoding="utf-8-sig")

    # Also write formatted XLSX with blue highlight for multi-line orders
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "shipping"
        ws.append(list(sorted_df.columns))
        blue = PatternFill("solid", fgColor="CCE5FF")
        # Mark as blue the rows whose order id is in multi_ids
        for _, row in sorted_df.iterrows():
            ws.append(list(row.values))
            if str(row["order id"]) in multi_ids:
                for cell in ws[ws.max_row]:
                    cell.fill = blue
        xlsx_name = out_csv.rsplit(".",1)[0] + "_formatted.xlsx"
        wb.save(xlsx_name)
        print(f"Wrote formatted workbook: {xlsx_name}")
    except Exception as e:
        print(f"Skipping XLSX formatting output due to: {e}")

    print(f"Done. Wrote: {out_csv}")

if __name__ == "__main__":
    main()
